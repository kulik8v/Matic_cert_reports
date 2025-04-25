#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
4_izvedeno.py

1. Читает journal.xlsx из <root>/templates и строит маппинг:
   Certificate -> Source File (из столбца A).

2. Для каждого сертификата:
   - копирует шаблон izvedeno_template.xlsx в <root>/Output/Izvedeno/<Certificate>.xlsx (если ещё нет)
   - на листах 'K_03_AB radovi', 'K_04_Armiracki' и 'K_00_REKAP' ищет теги:
       * '[data]' на листах radovi и Armiracki
       * '[extra_hours]' на листе K_00_REKAP
   - для найденных значений:
       – десятичные точки меняются на запятые
       – каждое найденное значение из источника используется только один раз;
         если уже использовалось — ставим 0
   - сохраняет итоговый файл

Скрипт лежит в папке "scripts" внутри корня проекта и может запускаться из любой директории.
"""
import sys
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

def main():
    # Определяем корень проекта из расположения скрипта
    scripts_dir = Path(__file__).resolve().parent
    root = scripts_dir.parent

    templates = root / 'templates'
    journal_fp = templates / 'journal.xlsx'
    template_fp = templates / 'izvedeno_template.xlsx'
    input_dir = root / 'Input'
    output_dir = root / 'Output' / 'Izvedeno'

    # Проверки существования необходимых путей
    for p in (journal_fp, template_fp, input_dir):
        if not p.exists():
            print(f"❌ Не найден: {p}", file=sys.stderr)
            sys.exit(1)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Читаем журнал и формируем маппинг Certificate -> Source File
    df = pd.read_excel(journal_fp, dtype=str)
    if 'Certificate' not in df.columns:
        print(f"❌ В '{journal_fp.name}' нет колонки 'Certificate'", file=sys.stderr)
        sys.exit(1)
    source_col = df.columns[0]
    mapping = {
        str(row['Certificate']).strip().removesuffix('.xlsx'): str(row[source_col]).strip().removesuffix('.xlsx')
        for _, row in df.dropna(subset=['Certificate']).iterrows()
    }

    # Обработка каждого сертификата
    for cert_name, src_name in mapping.items():
        target_fp = output_dir / f"{cert_name}.xlsx"
        source_fp = input_dir / f"{src_name}.xlsx"
        process_certificate(template_fp, target_fp, source_fp)


def process_certificate(template_file: Path, target_file: Path, source_path: Path):
    used_values = set()

    # Копируем шаблон, если нужно
    if not target_file.exists():
        shutil.copy(template_file, target_file)
        print(f"[CREATED] {target_file.name}")
    else:
        print(f"[EXISTS]  {target_file.name}")

    wb = load_workbook(target_file, data_only=False)
    wb_src = None
    if source_path.exists():
        wb_src = load_workbook(source_path, data_only=True)
    else:
        print(f"⚠ Источник не найден: {source_path}", file=sys.stderr)

    # Шаблонные имена листов
    sheets_data = ['K_03_AB radovi', 'K_04_Armiracki']
    for sheet_name in sheets_data:
        if sheet_name not in wb.sheetnames:
            print(f"⚠ Лист '{sheet_name}' не найден в {target_file.name}", file=sys.stderr)
            continue
        sheet = wb[sheet_name]
        print(f"→ Обработка листа '{sheet_name}'")

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '[data]' in cell.value:
                    key = sheet.cell(row=cell.row, column=2).value
                    found = None
                    if wb_src:
                        for sht in wb_src.worksheets:
                            for r in sht.iter_rows():
                                for c in r:
                                    if key and c.value and key in str(c.value):
                                        found = sht.cell(row=c.row, column=5).value
                                        break
                                if found is not None:
                                    break
                            if found is not None:
                                break

                    # Уникальность замены
                    raw_str = str(found) if found is not None else None
                    if raw_str in used_values:
                        found = None
                    else:
                        used_values.add(raw_str)

                    replacement = '0' if found is None else str(found).replace('.', ',')
                    original = cell.value
                    if original.strip() == '[data]':
                        sheet.cell(row=cell.row, column=cell.column).value = replacement
                    else:
                        sheet.cell(row=cell.row, column=cell.column).value = original.replace('[data]', replacement)
                    print(f"    Row {cell.row}: key='{key}' → '{replacement}'")

    # Обработка K_00_REKAP для [extra_hours]
    rekap = 'K_00_REKAP'
    if rekap in wb.sheetnames:
        sheet = wb[rekap]
        print(f"→ Обработка листа '{rekap}'")
        target_cell = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == '[extra_hours]':
                    target_cell = cell
                    break
            if target_cell:
                break
        if target_cell and wb_src:
            found_extra = None
            search_text = 'Izvođenje radova po zahtevu Naručioca'
            for sht in wb_src.worksheets:
                for r in sht.iter_rows():
                    for c in r:
                        if c.value == search_text:
                            found_extra = sht.cell(row=c.row, column=5).value
                            break
                    if found_extra is not None:
                        break
                if found_extra is not None:
                    break
            replacement = '0' if found_extra is None else str(found_extra).replace('.', ',')
            sheet.cell(row=target_cell.row, column=target_cell.column).value = replacement
            print(f"    [extra_hours] → '{replacement}'")
        else:
            print(f"⚠ Тег '[extra_hours]' не найден или источник отсутствует", file=sys.stderr)
    else:
        print(f"⚠ Лист '{rekap}' не найден в {target_file.name}", file=sys.stderr)

    # Сохраняем файл
    wb.save(target_file)
    print(f"[SAVED]   {target_file.name}\n")

if __name__ == '__main__':
    main()
