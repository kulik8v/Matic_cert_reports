#!/usr/bin/env python3
"""
Обновляет в journal.xlsx колонку "Total Amount Din" значениями из ячейки D32
каждого соответствующего файла situacija в папке Output/Situacija.
Скрипт находится в папке scripts внутри корня проекта.
"""
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

def update_journal_total_amount_din():
    # Определяем корень проекта: скрипт всегда лежит в <root>/scripts
    scripts_dir = Path(__file__).resolve().parent
    root = scripts_dir.parent

    templates_dir = root / "templates"
    situacija_dir = root / "Output" / "Situacija"
    journal_fp = templates_dir / "journal.xlsx"

    if not journal_fp.exists():
        print(f"❌ Файл журнала не найден: {journal_fp}", file=sys.stderr)
        sys.exit(1)

    # Загружаем журнал
    df = pd.read_excel(journal_fp)

    # Добавляем колонку, если её нет
    if "Total Amount Din" not in df.columns:
        df["Total Amount Din"] = None

    # Обновляем значения
    for idx, row in df.iterrows():
        invoice_name = row.get("Invoice")
        report_fp = situacija_dir / invoice_name

        if not report_fp.exists():
            print(f"⚠ Report not found: {report_fp}")
            continue

        wb = load_workbook(report_fp, data_only=True)
        ws = wb.active
        df.at[idx, "Total Amount Din"] = ws["D32"].value

    # Сохраняем изменения
    df.to_excel(journal_fp, index=False)
    print(f"✅ Journal updated: {journal_fp}")

if __name__ == "__main__":
    update_journal_total_amount_din()
